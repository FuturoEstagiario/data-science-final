import os
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(BASE_DIR, 'output')
EXCEL_PATH = os.path.join(OUTPUT_DIR, 'relatorio.xlsx')
PDF_PATH   = os.path.join(OUTPUT_DIR, 'relatorio.pdf')
HTML_PATH  = os.path.join(OUTPUT_DIR, 'relatorio.html')

_CAT_FILLS = {
    'Budget':    PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'Mid-range': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
    'Premium':   PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
}


def _borda():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)


def _cabecalho_cell(ws, row, col, valor, fill_hex='2F5496'):
    cell = ws.cell(row=row, column=col, value=valor)
    cell.fill   = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type='solid')
    cell.font   = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = _borda()
    return cell


# ─────────────────────────────────────────────────────────────
# EXCEL  (padrão aula8/exercicio8.py)
# ─────────────────────────────────────────────────────────────
def exportar_excel(df, resultados, log=print):
    log('[EXCEL] Gerando relatorio.xlsx...')
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb = Workbook()

    # ── Folha 1: Dados completos ─────────────────────────────
    ws = wb.active
    ws.title = 'Dados'
    colunas = ['Brand', 'Model', 'RAM', 'Storage', 'Color', 'Free',
               'Preco_BRL', 'Categoria', 'Top_Vendedor']
    larguras = [14, 22, 8, 10, 12, 8, 14, 12, 14]

    for col, h in enumerate(colunas, 1):
        _cabecalho_cell(ws, 1, col, h)
        ws.column_dimensions[get_column_letter(col)].width = larguras[col - 1]

    brd = _borda()
    for i, row in enumerate(df[colunas].itertuples(index=False), 2):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = brd
            if colunas[j - 1] == 'Categoria':
                cell.fill = _CAT_FILLS.get(str(val), PatternFill())

    ws.freeze_panes = 'A2'

    # ── Folha 2: Resumo por marca + gráfico ──────────────────
    ws2 = wb.create_sheet('Resumo por Marca')
    resumo = (df.groupby('Brand')
                .agg(Total=('Brand', 'count'),
                     Preco_Medio=('Preco_BRL', 'mean'),
                     RAM_Media=('RAM', 'mean'))
                .reset_index()
                .sort_values('Total', ascending=False)
                .head(15))
    resumo['Preco_Medio'] = resumo['Preco_Medio'].round(2)
    resumo['RAM_Media']   = resumo['RAM_Media'].round(1)

    hdrs2 = ['Marca', 'Total Modelos', 'Preço Médio (R$)', 'RAM Média (GB)']
    for col, h in enumerate(hdrs2, 1):
        _cabecalho_cell(ws2, 1, col, h)
        ws2.column_dimensions[get_column_letter(col)].width = 20

    for i, row in enumerate(resumo.itertuples(index=False), 2):
        for j, val in enumerate(row, 1):
            ws2.cell(row=i, column=j, value=val).border = brd

    # Gráfico de barras — padrão direto de aula8/exercicio8.py
    chart = BarChart()
    chart.type    = 'col'
    chart.title   = 'Total de Modelos por Marca (Top 15)'
    chart.y_axis.title = 'Modelos'
    chart.x_axis.title = 'Marca'
    chart.style   = 10
    chart.width   = 25
    chart.height  = 15
    n = len(resumo) + 1
    chart.add_data(Reference(ws2, min_col=2, min_row=1, max_row=n), titles_from_data=True)
    chart.set_categories(Reference(ws2, min_col=1, min_row=2, max_row=n))
    ws2.add_chart(chart, 'F2')

    # ── Folha 3: Previsões ML ────────────────────────────────
    ws3 = wb.create_sheet('Previsões ML')
    ws3.cell(row=1, column=1,
             value=f'Acurácia CV 5-fold: {resultados["acc_cv"]:.2%}').font = Font(bold=True)
    ws3.cell(row=2, column=1,
             value=f'Acurácia Teste: {resultados["acc_teste"]:.2%}').font = Font(bold=True)

    hdrs3 = ['Marca', 'Modelo', 'RAM', 'Storage', 'Preço R$', 'Categoria Real', 'Previsão', 'Acertou?']
    for col, h in enumerate(hdrs3, 1):
        _cabecalho_cell(ws3, 4, col, h, fill_hex='4472C4')
        ws3.column_dimensions[get_column_letter(col)].width = 16

    acerto_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    erro_fill   = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    df_test   = resultados['df_test']
    y_test    = resultados['y_test']
    previsoes = resultados['previsoes']

    for i, (_, fila) in enumerate(df_test.iterrows()):
        r = i + 5
        ws3.cell(row=r, column=1, value=fila['Brand']).border        = brd
        ws3.cell(row=r, column=2, value=fila['Model']).border        = brd
        ws3.cell(row=r, column=3, value=fila['RAM']).border          = brd
        ws3.cell(row=r, column=4, value=fila['Storage']).border      = brd
        ws3.cell(row=r, column=5, value=fila['Preco_BRL']).border    = brd
        ws3.cell(row=r, column=6, value=y_test[i]).border            = brd
        ws3.cell(row=r, column=7, value=previsoes[i]).border         = brd
        acertou = 'Sim' if y_test[i] == previsoes[i] else 'Não'
        cell_a  = ws3.cell(row=r, column=8, value=acertou)
        cell_a.border = brd
        cell_a.fill   = acerto_fill if acertou == 'Sim' else erro_fill

    wb.save(EXCEL_PATH)
    log(f'[EXCEL] Salvo em: {EXCEL_PATH}')


# ─────────────────────────────────────────────────────────────
# PDF  (padrão atv9/exercicio2_holerite.py)
# ─────────────────────────────────────────────────────────────
def exportar_pdf(df, resultados, taxa, log=print):
    log('[PDF] Gerando relatorio.pdf...')
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    doc = SimpleDocTemplate(
        PDF_PATH, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm
    )

    e_titulo    = ParagraphStyle('Titulo', fontName='Helvetica-Bold', fontSize=16,
                                  alignment=TA_CENTER, textColor=colors.HexColor('#003366'), spaceAfter=8)
    e_subtitulo = ParagraphStyle('Sub', fontName='Helvetica', fontSize=10,
                                  alignment=TA_CENTER, textColor=colors.HexColor('#555555'), spaceAfter=16)
    e_secao     = ParagraphStyle('Secao', fontName='Helvetica-Bold', fontSize=12,
                                  alignment=TA_LEFT, textColor=colors.HexColor('#003366'),
                                  spaceBefore=12, spaceAfter=6)
    e_rodape    = ParagraphStyle('Rodape', fontName='Helvetica', fontSize=8,
                                  alignment=TA_CENTER, textColor=colors.HexColor('#888888'), spaceBefore=20)

    elementos = []

    # Cabeçalho
    elementos.append(Paragraph('ANÁLISE DE MERCADO — SMARTPHONES', e_titulo))
    elementos.append(Paragraph(
        f'Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")} | Taxa EUR/BRL: R$ {taxa:.4f}',
        e_subtitulo
    ))

    # Resumo executivo
    elementos.append(Paragraph('RESUMO EXECUTIVO', e_secao))
    dados_resumo = [
        ['Indicador', 'Valor'],
        ['Total de Smartphones', str(len(df))],
        ['Marcas Analisadas', str(df['Brand'].nunique())],
        ['Preço Médio (R$)',   f'R$ {df["Preco_BRL"].mean():,.2f}'],
        ['Preço Mínimo (R$)',  f'R$ {df["Preco_BRL"].min():,.2f}'],
        ['Preço Máximo (R$)',  f'R$ {df["Preco_BRL"].max():,.2f}'],
        ['Acurácia ML (CV)',   f'{resultados["acc_cv"]:.2%}'],
        ['Acurácia ML (Teste)', f'{resultados["acc_teste"]:.2%}'],
    ]
    _adicionar_tabela(elementos, dados_resumo, [9*cm, 7*cm], '#003366')

    # Top 10 por preço
    elementos.append(Spacer(1, 0.4*cm))
    elementos.append(Paragraph('TOP 10 SMARTPHONES (Maior Preço)', e_secao))
    top10 = df.nlargest(10, 'Preco_BRL')
    dados_top = [['Marca', 'Modelo', 'RAM', 'Storage', 'Preço R$', 'Categoria']]
    for _, row in top10.iterrows():
        dados_top.append([
            row['Brand'],
            str(row['Model'])[:22],
            f'{int(row["RAM"])} GB',
            f'{int(row["Storage"])} GB',
            f'R$ {row["Preco_BRL"]:,.2f}',
            row['Categoria'],
        ])
    _adicionar_tabela(elementos, dados_top, [3*cm, 4.5*cm, 1.8*cm, 1.8*cm, 3.5*cm, 2.4*cm], '#2E7D32')

    # Distribuição por categoria
    elementos.append(Spacer(1, 0.4*cm))
    elementos.append(Paragraph('DISTRIBUIÇÃO POR CATEGORIA', e_secao))
    total = len(df)
    dados_cat = [['Categoria', 'Total', '% do Mercado']]
    for cat, cnt in df['Categoria'].value_counts().items():
        dados_cat.append([cat, str(cnt), f'{cnt/total*100:.1f}%'])
    _adicionar_tabela(elementos, dados_cat, [6*cm, 4*cm, 4*cm], '#C62828')

    # Rodapé
    elementos.append(Spacer(1, 1*cm))
    elementos.append(Paragraph(
        'Fontes: Dataset Kaggle · open.er-api.com (câmbio EUR→BRL) · Wikipedia | Projeto Acadêmico — Data Science',
        e_rodape
    ))

    doc.build(elementos)
    log(f'[PDF] Salvo em: {PDF_PATH}')


def _adicionar_tabela(elementos, dados, col_widths, header_hex):
    n_rows = len(dados)
    tab = Table(dados, colWidths=col_widths)
    style = [
        ('BACKGROUND',  (0, 0), (-1, 0), colors.HexColor(header_hex)),
        ('TEXTCOLOR',   (0, 0), (-1, 0), colors.white),
        ('FONTNAME',    (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME',    (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',    (0, 0), (-1, -1), 9),
        ('GRID',        (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('BOX',         (0, 0), (-1, -1), 1,   colors.HexColor(header_hex)),
        ('TOPPADDING',  (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ]
    # Linhas alternadas (manual — reportlab não tem ROWBACKGROUNDS)
    for i in range(1, n_rows):
        if i % 2 == 0:
            style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F0F4F8')))
    tab.setStyle(TableStyle(style))
    elementos.append(tab)


# ─────────────────────────────────────────────────────────────
# HTML
# ─────────────────────────────────────────────────────────────
def exportar_html(df, log=print):
    log('[HTML] Gerando relatorio.html...')
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    resumo = (df.groupby('Brand')
                .agg(Total=('Brand', 'count'), Preco_Medio=('Preco_BRL', 'mean'))
                .reset_index()
                .sort_values('Total', ascending=False)
                .head(10))

    linhas_resumo = ''.join(
        f'<tr><td>{r["Brand"]}</td><td>{r["Total"]}</td>'
        f'<td>R$ {r["Preco_Medio"]:,.2f}</td></tr>'
        for _, r in resumo.iterrows()
    )

    top20 = df.nlargest(20, 'Preco_BRL')
    cores = {'Budget': '#d4edda', 'Mid-range': '#fff3cd', 'Premium': '#f8d7da'}
    linhas_top = ''.join(
        f'<tr style="background:{cores.get(r["Categoria"], "#fff")}">'
        f'<td>{r["Brand"]}</td><td>{r["Model"]}</td>'
        f'<td>{int(r["RAM"])} GB</td><td>{int(r["Storage"])} GB</td>'
        f'<td>R$ {r["Preco_BRL"]:,.2f}</td><td>{r["Categoria"]}</td></tr>'
        for _, r in top20.iterrows()
    )

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<title>Análise de Mercado — Smartphones</title>
<style>
  body  {{ font-family: Arial, sans-serif; margin: 40px; background: #f5f5f5; color: #333; }}
  h1   {{ color: #003366; }}
  h2   {{ color: #2F5496; border-bottom: 2px solid #2F5496; padding-bottom: 5px; }}
  table {{ border-collapse: collapse; width: 100%; margin-bottom: 30px; background: white;
           box-shadow: 0 1px 4px rgba(0,0,0,.1); }}
  th   {{ background: #003366; color: white; padding: 10px 12px; text-align: left; }}
  td   {{ padding: 8px 12px; border: 1px solid #ddd; }}
  tr:hover {{ filter: brightness(0.97); }}
  footer {{ color: #888; font-size: 12px; margin-top: 30px; }}
</style>
</head>
<body>
<h1>📱 Análise de Mercado — Smartphones</h1>
<p>Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")} &nbsp;|&nbsp; {len(df)} smartphones analisados</p>

<h2>Resumo por Marca (Top 10)</h2>
<table>
  <tr><th>Marca</th><th>Total Modelos</th><th>Preço Médio</th></tr>
  {linhas_resumo}
</table>

<h2>Top 20 Smartphones — Maior Preço (R$)</h2>
<table>
  <tr><th>Marca</th><th>Modelo</th><th>RAM</th><th>Storage</th><th>Preço</th><th>Categoria</th></tr>
  {linhas_top}
</table>

<footer>
  Fontes: Dataset Kaggle &middot; open.er-api.com (câmbio EUR→BRL ao vivo) &middot; Wikipedia
  &nbsp;|&nbsp; Projeto Acadêmico — Data Science
</footer>
</body>
</html>"""

    with open(HTML_PATH, 'w', encoding='utf-8') as f:
        f.write(html)
    log(f'[HTML] Salvo em: {HTML_PATH}')


# ─────────────────────────────────────────────────────────────
# DASHBOARD DATA (view/js/data.js)
# ─────────────────────────────────────────────────────────────
def exportar_dashboard_data(df, resultados, taxa, log=print):
    log('[DASHBOARD] Exportando data.js para o dashboard...')
    view_js = os.path.join(BASE_DIR, 'view', 'js')
    os.makedirs(view_js, exist_ok=True)

    df = df.copy()

    by_brand = (df.groupby('Brand')
                  .agg(count=('Brand', 'count'),
                       avg_price=('Preco_BRL', 'mean'),
                       min_price=('Preco_BRL', 'min'),
                       max_price=('Preco_BRL', 'max'),
                       avg_ram=('RAM', 'mean'))
                  .reset_index()
                  .sort_values('count', ascending=False)
                  .head(15))

    cat_counts = df['Categoria'].astype(str).value_counts()
    total      = len(df)

    by_storage = (df.groupby('Storage').size()
                    .reset_index(name='count')
                    .sort_values('Storage'))

    by_ram = (df.groupby('RAM')
                .agg(count=('RAM', 'count'), avg_price=('Preco_BRL', 'mean'))
                .reset_index().sort_values('RAM'))

    phones = (df[['Brand', 'Model', 'RAM', 'Storage', 'Color', 'Preco_BRL', 'Categoria', 'Top_Vendedor']]
               .sort_values('Preco_BRL', ascending=False))

    top_brand = str(df['Brand'].value_counts().idxmax())

    data = {
        'meta': {
            'total':           int(total),
            'taxa_brl':        round(float(taxa), 4),
            'avg_price':       round(float(df['Preco_BRL'].mean()), 2),
            'min_price':       round(float(df['Preco_BRL'].min()), 2),
            'max_price':       round(float(df['Preco_BRL'].max()), 2),
            'marcas':          int(df['Brand'].nunique()),
            'top_brand':       top_brand,
            'top_brand_count': int(df['Brand'].value_counts().max()),
            'acc_cv':          round(float(resultados['acc_cv']), 4),
            'acc_teste':       round(float(resultados['acc_teste']), 4),
            'gerado_em':       datetime.now().strftime('%d/%m/%Y %H:%M'),
        },
        'by_brand': [
            {
                'brand':     str(row['Brand']),
                'count':     int(row['count']),
                'avg_price': round(float(row['avg_price']), 2),
                'min_price': round(float(row['min_price']), 2),
                'max_price': round(float(row['max_price']), 2),
                'avg_ram':   round(float(row['avg_ram']), 1),
            }
            for _, row in by_brand.iterrows()
        ],
        'by_category': [
            {'category': cat, 'count': int(cnt), 'pct': round(cnt / total * 100, 1)}
            for cat, cnt in cat_counts.items()
        ],
        'by_storage': [
            {'storage': int(row['Storage']), 'count': int(row['count'])}
            for _, row in by_storage.iterrows()
        ],
        'by_ram': [
            {'ram': int(row['RAM']), 'count': int(row['count']),
             'avg_price': round(float(row['avg_price']), 2)}
            for _, row in by_ram.iterrows()
        ],
        'phones': [
            {
                'brand':      str(row['Brand']),
                'model':      str(row['Model']),
                'ram':        int(row['RAM']),
                'storage':    int(row['Storage']),
                'color':      str(row['Color']),
                'price':      round(float(row['Preco_BRL']), 2),
                'category':   str(row['Categoria']),
                'top_seller': int(row['Top_Vendedor']),
            }
            for _, row in phones.iterrows()
        ],
    }

    out = os.path.join(view_js, 'data.js')
    with open(out, 'w', encoding='utf-8') as f:
        f.write('const DASHBOARD_DATA = ')
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write(';\n')

    log(f'[DASHBOARD] data.js gerado: {len(data["phones"])} phones | {out}')
